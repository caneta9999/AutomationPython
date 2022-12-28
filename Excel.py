#Credits and Thanks
#Dataset link: https://www.kaggle.com/datasets/marahim20/sales-dataset-of-supermarket

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from datetime import datetime

def create_pivot_table(dataset, columnsArray, index, columns, values, aggfunc, output="pivot_table.xlsx", sheetName="Report", startRow=5):
    df = ""
    if ".csv" in dataset:
        df = pd.read_csv(dataset)
    elif ".xlsx" in dataset:
        df = pd.read_excel(dataset)
    df = df[columnsArray]
    pivot_table = df.pivot_table(index=index,columns=columns,values=values, aggfunc=aggfunc).round(0)
    pivot_table.to_excel(output,sheetName,startrow=startRow)

def read_pivot_table(pivot_table="pivot_table.xlsx", sheetName="Report"):
    wb = load_workbook(f'./{pivot_table}')
    sheet = wb[sheetName]
    min_col = wb.active.min_column
    max_col = wb.active.max_column
    min_row = wb.active.min_row
    max_row = wb.active.max_row
    return {"wb": wb,"sheet": sheet, "min_col": min_col, "max_col": max_col, "min_row":min_row, "max_row": max_row}

def locate(read):
    data = Reference(read['sheet'],min_col=read['min_col']+1,max_col=read['max_col'],min_row=read['min_row'],max_row=read['max_row'])
    categories = Reference(read['sheet'],min_col=read['min_col'],max_col=read['min_col'],min_row=read['min_row']+1,max_row=read['max_row']) 
    return {"data": data, "categories": categories}

def create_chart(read,title="Sales", position="B15"):
    barchart = BarChart()
    location = locate(read)
    barchart.add_data(location['data'], titles_from_data=True)
    barchart.set_categories(location['categories'])
    read['sheet'].add_chart(barchart, position)
    barchart.title = title
    barchart.style = 5

def write_formulas(read,style="Currency",font="Calibri",font_size=7):
    for i in range(read['min_col']+1, read['max_col']+1):
        letter = get_column_letter(i)
        read['sheet'][f'{letter}{read["max_row"] + 1}'] = f'=SUM({letter}{read["min_row"] + 1}:{letter}{read["max_row"]})'
        read['sheet'][f'{letter}{read["max_row"] + 1}'].style = style
        read['sheet'][f'{letter}{read["max_row"] + 1}'].font = Font(font, size=font_size)

def save_report(read, name="report"):
    create_chart(read)
    write_formulas(read)
    read['wb'].save(f'./{name}_{datetime.now().strftime("%B%Y").lower()}.xlsx')
    
columnsArray = ['Gender','Category','Total']
create_pivot_table("sales.csv",columnsArray,columnsArray[0],columnsArray[1],columnsArray[2],'sum')
save_report(read_pivot_table())
